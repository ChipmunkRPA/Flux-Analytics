#!/usr/bin/env python3
"""
summarize_amount_by_period.py

Read an Excel file, summarize the `Amount` column by monthly `Period`,
compute month-over-month flux and percent change, save a summary Excel,
and create a month-over-month chart (Amount and MoM flux).
"""
import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

import os
import pandas as pd


def _read_table_file(input_path, sheet=0):
    """Read a table from CSV or Excel based on file extension.
    - .csv -> pandas.read_csv
    - .xlsx/.xls/.xlsm -> pandas.read_excel(sheet_name=sheet)
    - otherwise: try Excel first, then CSV
    """
    path_lower = str(input_path).lower()
    try:
        if path_lower.endswith('.csv'):
            try:
                df = pd.read_csv(input_path)
            except Exception:
                # Fallback encodings
                for enc in ('utf-8-sig', 'utf-8', 'latin1'):
                    try:
                        df = pd.read_csv(input_path, encoding=enc)
                        break
                    except Exception:
                        continue
                # re-raise if all fail
                df = pd.read_csv(input_path)
            # strip column name whitespace
            try:
                df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
            except Exception:
                pass
            return df
        if path_lower.endswith('.xlsx') or path_lower.endswith('.xls') or path_lower.endswith('.xlsm'):
            df = pd.read_excel(input_path, sheet_name=sheet)
            try:
                df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
            except Exception:
                pass
            return df
        # Unknown extension: attempt Excel then CSV
        try:
            df = pd.read_excel(input_path, sheet_name=sheet)
            try:
                df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
            except Exception:
                pass
            return df
        except Exception:
            try:
                df = pd.read_csv(input_path)
            except Exception:
                for enc in ('utf-8-sig', 'utf-8', 'latin1'):
                    try:
                        df = pd.read_csv(input_path, encoding=enc)
                        break
                    except Exception:
                        continue
                df = pd.read_csv(input_path)
            try:
                df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
            except Exception:
                pass
            return df
    except Exception as e:
        raise e


def _coerce_amount_inplace(df, value_col):
    """Ensure df[value_col] is numeric, handling $, commas, and parentheses.
    Modifies df in place.
    """
    try:
        if value_col in df.columns:
            if not pd.api.types.is_numeric_dtype(df[value_col]):
                s = df[value_col].astype(str).str.strip()
                # Convert parentheses negatives and remove currency/commas
                s = s.str.replace('(', '-', regex=False).str.replace(')', '', regex=False)
                s = s.str.replace(',', '', regex=False).str.replace('$', '', regex=False)
                df[value_col] = pd.to_numeric(s, errors='coerce')
    except Exception:
        pass


def ensure_openai_env():
    """Ensure `.env` exists and contains OPENAI_API_KEY (and OPENAI_BASE_URL).
    If missing, prompt the user, write/update `.env`, and set os.environ.
    """
    env_path = os.path.join(os.getcwd(), '.env')
    existing = {}
    # Load existing .env key/values if present
    if os.path.exists(env_path):
        try:
            with open(env_path, 'r', encoding='utf-8') as f:
                for ln in f:
                    ln = ln.strip()
                    if not ln or ln.startswith('#') or '=' not in ln:
                        continue
                    k, v = ln.split('=', 1)
                    k = k.strip()
                    v = v.strip().strip('"').strip("'")
                    if k:
                        existing[k] = v
        except Exception:
            try:
                with open(env_path, 'r') as f:
                    for ln in f:
                        ln = ln.strip()
                        if not ln or ln.startswith('#') or '=' not in ln:
                            continue
                        k, v = ln.split('=', 1)
                        k = k.strip()
                        v = v.strip().strip('"').strip("'")
                        if k:
                            existing[k] = v
            except Exception:
                pass

    has_api_key = bool(os.environ.get('OPENAI_API_KEY') or existing.get('OPENAI_API_KEY'))

    # If we already have a key (either env or .env), ensure env is populated and return
    if has_api_key and os.path.exists(env_path):
        if 'OPENAI_API_KEY' not in os.environ and existing.get('OPENAI_API_KEY'):
            os.environ['OPENAI_API_KEY'] = existing['OPENAI_API_KEY']
        if existing.get('OPENAI_BASE_URL') and 'OPENAI_BASE_URL' not in os.environ:
            os.environ['OPENAI_BASE_URL'] = existing['OPENAI_BASE_URL']
        return

    print('OpenAI API key not configured. Let\'s set it up now.')
    # Prompt user for API key (shown as you type)
    try:
        api_key = input('Enter your OpenAI API key (will be shown): ').strip()
    except EOFError:
        api_key = ''
    # Display back what was entered so user can verify
    if api_key:
        print(f'You entered OPENAI_API_KEY: {api_key}')

    default_base = 'https://api.openai.com/v1'
    try:
        base_url_input = input(f'Enter OpenAI base URL. Hit enter to use default [{default_base}]: ').strip()
    except EOFError:
        base_url_input = ''
    base_url = base_url_input or default_base

    # Update map and environment
    if api_key:
        existing['OPENAI_API_KEY'] = api_key
        os.environ['OPENAI_API_KEY'] = api_key
    if base_url:
        existing['OPENAI_BASE_URL'] = base_url
        os.environ['OPENAI_BASE_URL'] = base_url

    # Write or update .env
    try:
        with open(env_path, 'w', encoding='utf-8') as f:
            for k, v in existing.items():
                if not k:
                    continue
                f.write(f'{k}={v}\n')
        print(f'Saved OpenAI settings to {env_path}')
    except Exception:
        try:
            with open(env_path, 'w') as f:
                for k, v in existing.items():
                    if not k:
                        continue
                    f.write(f'{k}={v}\n')
            print(f'Saved OpenAI settings to {env_path}')
        except Exception as e:
            print('Failed to write .env:', e)


def summarize(input_file, sheet=0, date_col="Period", value_col="Amount", output_excel="summary_flux.xlsx", include_department=False, department_col="Department", include_class=False, class_col="Class", fy_end_month=1):
    df = _read_table_file(input_file, sheet=sheet)
    # Normalize common column names (case/whitespace)
    lowered = {str(c).strip().lower(): c for c in df.columns}
    # Remap requested date/value columns if their case/spacing differs
    if date_col.lower() in lowered and lowered[date_col.lower()] != date_col:
        date_col = lowered[date_col.lower()]
    if value_col.lower() in lowered and lowered[value_col.lower()] != value_col:
        value_col = lowered[value_col.lower()]
    # Coerce Amount-like column to numeric for CSV inputs
    _coerce_amount_inplace(df, value_col)

    # Enforce required columns: Period (date_col), Amount (value_col), and Memo-like column
    missing = []
    if date_col not in df.columns:
        missing.append(date_col)
    if value_col not in df.columns:
        missing.append(value_col)
    if include_department and department_col not in df.columns:
        missing.append(department_col)
    if include_class and class_col not in df.columns:
        missing.append(class_col)

    # Detect memo/description column (required)
    memo_col_required = None
    possible_memo_names = ['Memo', 'memo', 'MEMO', 'Description', 'description', 'DESC', 'Details', 'details', 'Notes', 'notes']
    for c in df.columns:
        if c in possible_memo_names:
            memo_col_required = c
            break
    if memo_col_required is None:
        lowered = {c.lower(): c for c in df.columns}
        for name in possible_memo_names:
            if name.lower() in lowered:
                memo_col_required = lowered[name.lower()]
                break
    if memo_col_required is None:
        missing.append('Memo (or Description/Details/Notes)')

    if missing:
        raise ValueError(f"Required column(s) missing: {', '.join(missing)}. Available columns: {list(df.columns)}")
    # At this point required columns were validated above

    # Parse dates (coerce invalids)
    original_period_values = df[date_col].copy()  # Keep original values for custom parsing
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    if df[date_col].isna().all():
        # Try parsing as Period (like '2025-01' or '202501')
        try:
            df[date_col] = pd.PeriodIndex(original_period_values.astype(str), freq='M').to_timestamp()
        except Exception:
            # Try parsing month-year format like 'Feb-25'
            try:
                # Convert various date formats to proper date
                def parse_month_year(val):
                    if pd.isna(val):
                        return pd.NaT
                    val_str = str(val).strip()
                    if '-' in val_str and len(val_str.split('-')) == 2:
                        part1, part2 = val_str.split('-')
                        # Convert month name to number
                        month_map = {
                            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
                        }
                        
                        # Try both formats: "Feb-25" and "25-Feb" (both are month-year)
                        month_num = month_map.get(part1.lower())
                        if month_num:
                            # Format: "Feb-25" (month-year)
                            year = int('20' + part2) if len(part2) == 2 else int(part2)
                            return pd.Timestamp(year=year, month=month_num, day=1)
                        else:
                            month_num = month_map.get(part2.lower())
                            if month_num:
                                # Format: "25-Feb" (year-month) - 25 means 2025
                                year = int('20' + part1) if len(part1) == 2 else int(part1)
                                return pd.Timestamp(year=year, month=month_num, day=1)
                    return pd.NaT
                
                df[date_col] = original_period_values.apply(parse_month_year)
                if df[date_col].isna().all():
                    raise ValueError(f"Could not parse '{date_col}' as dates or periods.")
            except Exception:
                raise ValueError(f"Could not parse '{date_col}' as dates or periods.")

    # Create monthly period (timestamp at month start)
    df['PeriodMonth'] = df[date_col].dt.to_period('M').dt.to_timestamp()

    # Attach fiscal quarter start and label to each row for later per-quarter sheets
    def _fiscal_quarter_info_row(ts):
        # ts is a Timestamp at month start
        m = int(ts.month)
        y = int(ts.year)
        try:
            fy_end = int(fy_end_month)
        except Exception:
            fy_end = 1
        if fy_end < 1 or fy_end > 12:
            fy_end = 1
        # Fiscal year label uses the year of the fiscal year end month
        fy_label_end = y if m <= fy_end else y + 1
        # Q1 starts the month after the end month
        q1_start = 1 if fy_end == 12 else fy_end + 1
        months_since_q1 = (m - q1_start) % 12
        q = (months_since_q1 // 3) + 1
        start_month = ((q1_start + (q - 1) * 3 - 1) % 12) + 1
        start_year = y if start_month <= m else y - 1
        return pd.Timestamp(year=start_year, month=start_month, day=1), f"FY{fy_label_end}-Q{q}"

    df[['FiscalQuarterStart', 'FiscalQuarterLabel']] = df['PeriodMonth'].apply(lambda ts: pd.Series(list(_fiscal_quarter_info_row(ts))))

    # Amount existence validated above

    # Aggregate
    monthly = df.groupby('PeriodMonth', sort=True)[value_col].sum().rename('Amount')
    flux = monthly.diff().fillna(0).rename('MoM_Flux')
    pct = monthly.pct_change().fillna(0).rename('MoM_Pct') * 100

    summary = pd.concat([monthly, flux, pct], axis=1)

    # Fiscal-quarter aggregation (dynamic fiscal year end month)
    def _fiscal_quarter_info(ts: pd.Timestamp):
        m = int(ts.month)
        y = int(ts.year)
        try:
            fy_end = int(fy_end_month)
        except Exception:
            fy_end = 1
        if fy_end < 1 or fy_end > 12:
            fy_end = 1
        # Fiscal year label uses the year of the fiscal year end month
        fy_label_end = y if m <= fy_end else y + 1
        # Q1 starts the month after the end month
        q1_start = 1 if fy_end == 12 else fy_end + 1
        months_since_q1 = (m - q1_start) % 12
        q = (months_since_q1 // 3) + 1
        start_month = ((q1_start + (q - 1) * 3 - 1) % 12) + 1
        start_year = y if start_month <= m else y - 1
        quarter_label = f"FY{fy_label_end}-Q{q}"
        quarter_start = pd.Timestamp(year=start_year, month=start_month, day=1)
        return quarter_start, quarter_label

    # Build quarterly series by mapping each month to its fiscal quarter
    q_map = {}
    for ts, amt in monthly.items():
        qs, qlabel = _fiscal_quarter_info(ts)
        if qs in q_map:
            q_map[qs]['Amount'] += float(amt)
        else:
            q_map[qs] = {'Amount': float(amt), 'QuarterLabel': qlabel}

    if q_map:
        q_items = sorted(q_map.items(), key=lambda x: x[0])
        quarter_index = [k for k, v in q_items]
        quarter_amounts = pd.Series([v['Amount'] for k, v in q_items], index=quarter_index).rename('Amount')
        q_flux = quarter_amounts.diff().fillna(0).rename('QoQ_Flux')
        q_pct = quarter_amounts.pct_change().fillna(0).rename('QoQ_Pct') * 100
        quarter_summary = pd.concat([quarter_amounts, q_flux, q_pct], axis=1)
        # add readable label column
        quarter_summary = quarter_summary.reset_index().rename(columns={'index': 'QuarterStart'})
        quarter_summary['Quarter'] = quarter_summary['QuarterStart'].apply(lambda ts: q_map.get(ts, {}).get('QuarterLabel', str(ts)))
        # order columns
        quarter_summary = quarter_summary[['Quarter', 'QuarterStart', 'Amount', 'QoQ_Flux', 'QoQ_Pct']]
    else:
        quarter_summary = None

    # Save to Excel: include a Summary sheet and one sheet per month with transactions
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # Summary sheet (trend) - write Period as YYYY-MM strings so Excel x-axis shows months
        summary_df = summary.reset_index()
        # Force Period column to YYYY-MM strings so Excel will use those as categorical x-axis labels
        try:
            summary_df.iloc[:, 0] = pd.to_datetime(summary_df.iloc[:, 0]).dt.strftime('%Y-%m')
        except Exception:
            # fallback: cast to string
            summary_df.iloc[:, 0] = summary_df.iloc[:, 0].astype(str)
        summary_df.rename(columns={summary_df.columns[0]: 'Period'}, inplace=True)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # write quarterly sheet if available
        if quarter_summary is not None:
            try:
                # format QuarterStart as YYYY-MM for readability
                q_out = quarter_summary.copy()
                q_out['QuarterStart'] = q_out['QuarterStart'].dt.strftime('%Y-%m')
                q_out.to_excel(writer, sheet_name='Quarterly', index=False)
            except Exception:
                pass

        # Per-month transaction sheets: filter original df for each month
        for period_ts, group in df.groupby('PeriodMonth'):
            # Sheet name like 'YYYY-MM'
            sheet_name = period_ts.strftime('%Y-%m')
            # write group's original columns (reset index)
            if 'PeriodMonth' in group.columns:
                group_sorted = group.sort_values(by=['PeriodMonth'], ignore_index=True)
            else:
                group_sorted = group.reset_index(drop=True)
            # Drop internal PeriodMonth column to keep original appearance
            group_to_write = group_sorted.drop(columns=['PeriodMonth'], errors='ignore')
            # Excel has a 31-char sheet name limit; ensure name fits
            sheet_name = sheet_name[:31]
            try:
                group_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception:
                # fallback: use a safe sheet name enumerated
                safe_name = sheet_name[:25] + '_' + str(abs(hash(sheet_name)) % 1000)
                group_to_write.to_excel(writer, sheet_name=safe_name, index=False)

            # Optionally, also create per-department sheets within each month
            if include_department and department_col in group.columns:
                try:
                    for dept_value, dept_group in group.groupby(department_col):
                        base = period_ts.strftime('%Y-%m')
                        dept_label = str(dept_value) if pd.notna(dept_value) else 'Unknown'
                        # Build a concise sheet name within Excel's 31 char limit
                        candidate = f"{base}-{dept_label}"
                        sheet_name_dept = candidate[:31]
                        dept_out = dept_group.drop(columns=['PeriodMonth'], errors='ignore').reset_index(drop=True)
                        try:
                            dept_out.to_excel(writer, sheet_name=sheet_name_dept, index=False)
                        except Exception:
                            # Fallback: hash to avoid duplicates/length issues
                            safe_name = (base + '-' + str(abs(hash(dept_label)) % 10000))[:31]
                            dept_out.to_excel(writer, sheet_name=safe_name, index=False)
                except Exception:
                    pass

            # Optionally, also create per-class sheets within each month
            if include_class and class_col in group.columns:
                try:
                    for class_value, class_group in group.groupby(class_col):
                        base = period_ts.strftime('%Y-%m')
                        class_label = str(class_value) if pd.notna(class_value) else 'Unknown'
                        candidate = f"{base}-{class_label}"
                        sheet_name_class = candidate[:31]
                        class_out = class_group.drop(columns=['PeriodMonth'], errors='ignore').reset_index(drop=True)
                        try:
                            class_out.to_excel(writer, sheet_name=sheet_name_class, index=False)
                        except Exception:
                            safe_name = (base + '-' + str(abs(hash(class_label)) % 10000))[:31]
                            class_out.to_excel(writer, sheet_name=safe_name, index=False)
                except Exception:
                    pass

        # Per-quarter transaction sheets: group by FiscalQuarterStart and write each quarter's transactions
        try:
            for qstart, qgroup in df.groupby('FiscalQuarterStart'):
                qlabel = None
                if 'FiscalQuarterLabel' in qgroup.columns:
                    qlabel = qgroup['FiscalQuarterLabel'].iloc[0]
                sheet_name = (str(qlabel) if qlabel is not None else qstart.strftime('%Y-%m'))
                sheet_name = sheet_name[:31]
                try:
                    q_to_write = qgroup.drop(columns=['PeriodMonth'], errors='ignore')
                    q_to_write = q_to_write.drop(columns=['FiscalQuarterStart', 'FiscalQuarterLabel'], errors='ignore')
                    q_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception:
                    safe_name = sheet_name[:25] + '_' + str(abs(hash(sheet_name)) % 1000)
                    q_to_write.to_excel(writer, sheet_name=safe_name, index=False)

                # Optionally, also create per-department sheets within each quarter
                if include_department and department_col in qgroup.columns:
                    try:
                        for dept_value, dept_group in qgroup.groupby(department_col):
                            base = (str(qlabel) if qlabel is not None else qstart.strftime('%Y-%m'))
                            dept_label = str(dept_value) if pd.notna(dept_value) else 'Unknown'
                            candidate = f"{base}-{dept_label}"
                            sheet_name_dept = candidate[:31]
                            dept_out = dept_group.drop(columns=['PeriodMonth', 'FiscalQuarterStart', 'FiscalQuarterLabel'], errors='ignore').reset_index(drop=True)
                            try:
                                dept_out.to_excel(writer, sheet_name=sheet_name_dept, index=False)
                            except Exception:
                                safe_name = (base + '-' + str(abs(hash(dept_label)) % 10000))[:31]
                                dept_out.to_excel(writer, sheet_name=safe_name, index=False)
                    except Exception:
                        pass

                # Optionally, also create per-class sheets within each quarter
                if include_class and class_col in qgroup.columns:
                    try:
                        for class_value, class_group in qgroup.groupby(class_col):
                            base = (str(qlabel) if qlabel is not None else qstart.strftime('%Y-%m'))
                            class_label = str(class_value) if pd.notna(class_value) else 'Unknown'
                            candidate = f"{base}-{class_label}"
                            sheet_name_class = candidate[:31]
                            class_out = class_group.drop(columns=['PeriodMonth', 'FiscalQuarterStart', 'FiscalQuarterLabel'], errors='ignore').reset_index(drop=True)
                            try:
                                class_out.to_excel(writer, sheet_name=sheet_name_class, index=False)
                            except Exception:
                                safe_name = (base + '-' + str(abs(hash(class_label)) % 10000))[:31]
                                class_out.to_excel(writer, sheet_name=safe_name, index=False)
                    except Exception:
                        pass
        except Exception:
            pass

    # Chart generation and embedding removed by user request.

    return summary


def perform_openai_analysis(output_excel, df_all, summary_df, api_env_var='OPENAI_API_KEY', date_col='Period', value_col='Amount', model=None, dry_run=False, max_tokens=None, prompt_cap=0, analysis_mode='mom', include_department=False, department_col='Department', include_class=False, class_col='Class'):
    """Call OpenAI (if API key present in env) to analyze MoM fluctuations.

    This function prefers per-month sheets in `output_excel` (sheets named `YYYY-MM`) and
    analyzes all adjacent month pairs found there. If those sheets are missing it falls
    back to grouping `df_all` by `PeriodMonth`.
    """
    # If this is a dry run, skip loading .env and OpenAI imports/keys
    if not dry_run:
        # load .env if present
        env_path = os.path.join(os.getcwd(), '.env')
        if os.path.exists(env_path):
            try:
                # Prefer UTF-8 for .env; fall back to default if needed
                try:
                    with open(env_path, 'r', encoding='utf-8') as f:
                        for ln in f:
                            ln = ln.strip()
                            if not ln or ln.startswith('#'):
                                continue
                            if '=' not in ln:
                                continue
                            k, v = ln.split('=', 1)
                            k = k.strip()
                            v = v.strip().strip('"').strip("'")
                            if k not in os.environ:
                                os.environ[k] = v
                except Exception:
                    with open(env_path, 'r') as f:
                        for ln in f:
                            ln = ln.strip()
                            if not ln or ln.startswith('#'):
                                continue
                            if '=' not in ln:
                                continue
                            k, v = ln.split('=', 1)
                            k = k.strip()
                            v = v.strip().strip('"').strip("'")
                            if k not in os.environ:
                                os.environ[k] = v
            except Exception:
                pass

        key = os.environ.get(api_env_var)
        if not key:
            print(f"OPENAI API key not found in environment variable '{api_env_var}'. Skipping OpenAI analysis.")
            return None

        try:
            import openai
        except Exception:
            print("openai package not installed. Install with: pip install openai")
            return None

        openai.api_key = key

    # Normalize analysis mode
    try:
        mode = (analysis_mode or 'mom').strip().lower()
    except Exception:
        mode = 'mom'

    # Build initial prompt pieces (defer content based on mode)
    lines = []
    if mode == 'mom':
        lines.append("You are given monthly totals and transaction memo breakdowns. Explain the main month-over-month fluctuations and likely causes.")
        if include_department and department_col in df_all.columns:
            lines.append(f"The dataset includes a '{department_col}' column. Incorporate department-level context when attributing changes, and consider Department together with memo and Amount (analyze (Department, Memo, Amount) jointly).")
        if include_department and department_col not in df_all.columns:
            lines.append(f"Note: '{department_col}' column not found; skipping Department-level analysis.")
        if include_class and class_col in df_all.columns:
            lines.append(f"The dataset includes a '{class_col}' column. Incorporate class-level context when attributing changes, and consider Class together with memo and Amount (analyze (Class, Memo, Amount) jointly).")
        if include_class and class_col not in df_all.columns:
            lines.append(f"Note: '{class_col}' column not found; skipping Class-level analysis.")
        try:
            lines.append("Summary table (Period, Amount, MoM_Flux, MoM_Pct):")
            lines.append(summary_df.to_string(index=False))
        except Exception:
            pass
    elif mode == 'qoq':
        lines.append("You are given fiscal-quarter totals and (when available) quarter-level detail. Explain the main quarter-over-quarter fluctuations and likely causes.")
        if include_department and department_col in df_all.columns:
            lines.append(f"The dataset includes a '{department_col}' column. Incorporate department-level context when attributing changes, and consider Department together with memo and Amount (analyze (Department, Memo, Amount) jointly).")
        if include_department and department_col not in df_all.columns:
            lines.append(f"Note: '{department_col}' column not found; skipping Department-level analysis.")
        if include_class and class_col in df_all.columns:
            lines.append(f"The dataset includes a '{class_col}' column. Incorporate class-level context when attributing changes, and consider Class together with memo and Amount (analyze (Class, Memo, Amount) jointly).")
        if include_class and class_col not in df_all.columns:
            lines.append(f"Note: '{class_col}' column not found; skipping Class-level analysis.")
    else:
        # Fallback to MoM if unrecognized
        mode = 'mom'
        lines.append("You are given monthly totals and transaction memo breakdowns. Explain the main month-over-month fluctuations and likely causes.")
        try:
            lines.append("Summary table (Period, Amount, MoM_Flux, MoM_Pct):")
            lines.append(summary_df.to_string(index=False))
        except Exception:
            pass

    # detect memo column
    memo_col = None
    possible_memo_names = ['memo', 'Memo', 'MEMO', 'description', 'Description', 'DESC', 'Details', 'notes', 'Notes']
    for c in df_all.columns:
        if c in possible_memo_names:
            memo_col = c
            break
    if memo_col is None:
        lowered = {c.lower(): c for c in df_all.columns}
        for name in possible_memo_names:
            if name.lower() in lowered:
                memo_col = lowered[name.lower()]
                break

    # Initialize holders that may be populated depending on mode
    month_groups = {}
    months_sorted = []
    quarter_groups = {}
    if mode == 'mom':
        # read per-month sheets from output_excel if available
        try:
            xl = pd.ExcelFile(output_excel)
            import re
            month_sheet_names = [s for s in xl.sheet_names if re.match(r"^\d{4}-\d{2}$", s)]
            for s in month_sheet_names:
                try:
                    df_sheet = pd.read_excel(output_excel, sheet_name=s)
                    period_ts = pd.to_datetime(s + '-01')
                    df_sheet['PeriodMonth'] = period_ts
                    month_groups[period_ts] = df_sheet
                except Exception:
                    continue
            months_sorted = sorted(month_groups.keys())
        except Exception:
            months_sorted = []
    else:
        # For QoQ mode, we may still collect quarter groups from the workbook
        try:
            xl = pd.ExcelFile(output_excel)
            import re
            # fallback: also include sheet named 'Quarterly' to infer ordering
            if 'Quarterly' in xl.sheet_names:
                try:
                    qdf = pd.read_excel(output_excel, sheet_name='Quarterly')
                    if 'QuarterStart' in qdf.columns:
                        for _, r in qdf.iterrows():
                            try:
                                ts = pd.to_datetime(str(r['QuarterStart']) + '-01')
                                quarter_groups[ts] = None
                            except Exception:
                                continue
                except Exception:
                    pass
            for s in xl.sheet_names:
                if re.match(r"^FY\d{4}-Q[1-4]$", s):
                    try:
                        df_sheet = pd.read_excel(output_excel, sheet_name=s)
                        quarter_groups[s] = df_sheet
                    except Exception:
                        continue
        except Exception:
            quarter_groups = {}

    # fallback to grouping raw dataframe (MoM only)
    if mode == 'mom' and not months_sorted:
        if date_col not in df_all.columns:
            possible = [c for c in df_all.columns if 'period' in c.lower() or 'date' in c.lower()]
            if possible:
                date_col = possible[0]
        if date_col in df_all.columns:
            df_all[date_col] = pd.to_datetime(df_all[date_col], errors='coerce')
            df_all['PeriodMonth'] = df_all[date_col].dt.to_period('M').dt.to_timestamp()
        if 'PeriodMonth' in df_all.columns:
            month_groups = {period: grp for period, grp in df_all.groupby('PeriodMonth')}
            months_sorted = sorted(month_groups.keys())

    # Add an explicit, unambiguous mapping of adjacent month pairs to their MoM values
    # This prevents downstream consumers (or models) from mis-associating the MoM value
    # with the wrong month-pair when reading the summary table.
    if mode == 'mom':
        try:
            pair_summary_lines = []
            # helper to extract MoM values from the passed summary_df for a given period (timestamp)
            def _get_mom_values_for(period_ts):
                mom_flux = None
                mom_pct = None
                try:
                    # If summary_df has the period as its index
                    if hasattr(summary_df, 'index') and period_ts in summary_df.index:
                        row = summary_df.loc[period_ts]
                        mom_flux = row.get('MoM_Flux') if 'MoM_Flux' in summary_df.columns else None
                        mom_pct = row.get('MoM_Pct') if 'MoM_Pct' in summary_df.columns else None
                    else:
                        # try matching by YYYY-MM string in a 'Period' column or index string
                        key = period_ts.strftime('%Y-%m')
                        if 'Period' in summary_df.columns:
                            match = summary_df[summary_df['Period'].astype(str).str.startswith(key)]
                            if not match.empty:
                                mom_flux = match.iloc[0].get('MoM_Flux', None)
                                mom_pct = match.iloc[0].get('MoM_Pct', None)
                        else:
                            # attempt to find an index that starts with the key
                            for idx in summary_df.index:
                                if str(idx).startswith(key):
                                    row = summary_df.loc[idx]
                                    mom_flux = row.get('MoM_Flux') if 'MoM_Flux' in summary_df.columns else None
                                    mom_pct = row.get('MoM_Pct') if 'MoM_Pct' in summary_df.columns else None
                                    break
                except Exception:
                    pass
                return mom_flux, mom_pct

            if months_sorted and len(months_sorted) > 1:
                for i in range(1, len(months_sorted)):
                    m_prev = months_sorted[i-1]
                    m_curr = months_sorted[i]
                    mom_flux, mom_pct = _get_mom_values_for(m_curr)
                    if mom_flux is None or (isinstance(mom_flux, float) and pd.isna(mom_flux)):
                        pair_summary_lines.append(f"{m_prev.strftime('%Y-%m')} -> {m_curr.strftime('%Y-%m')}: MoM_Flux: N/A, MoM_Pct: N/A")
                    else:
                        # format numbers nicely
                        try:
                            mom_flux_f = f"{mom_flux:,.2f}"
                        except Exception:
                            mom_flux_f = str(mom_flux)
                        try:
                            mom_pct_f = f"{mom_pct:.1f}%" if mom_pct is not None and not pd.isna(mom_pct) else 'N/A'
                        except Exception:
                            mom_pct_f = str(mom_pct)
                        pair_summary_lines.append(f"{m_prev.strftime('%Y-%m')} -> {m_curr.strftime('%Y-%m')}: MoM_Flux: {mom_flux_f}, MoM_Pct: {mom_pct_f}")

            if pair_summary_lines:
                lines.append('\nExplicit adjacent-pair MoM values:')
                lines.append('\n'.join(pair_summary_lines))
        except Exception:
            # if anything goes wrong building the explicit mapping, continue without it
            pass

    # If a quarterly sheet was produced, attempt to read it and add QoQ explicit mapping (QoQ only)
    if mode == 'qoq':
        try:
            # try reading Quarterly from the output Excel if present
            q_summary = None
            try:
                qdf = pd.read_excel(output_excel, sheet_name='Quarterly')
                if not qdf.empty:
                    q_summary = qdf
            except Exception:
                # fallback: build from summary_df if possible
                if 'Quarter' in summary_df.columns and 'QoQ_Flux' in summary_df.columns:
                    q_summary = summary_df[['Quarter', 'QoQ_Flux', 'QoQ_Pct']]
            if q_summary is not None:
                lines.append('\nQuarterly summary (fiscal quarters):')
                # show the quarter table
                try:
                    lines.append(q_summary.to_string(index=False))
                except Exception:
                    lines.append(str(q_summary))
                # explicit adjacent QoQ mapping
                try:
                    q_lines = []
                    # expect a column named 'Quarter' or 'QuarterStart' to determine order
                    if 'QuarterStart' in q_summary.columns:
                        q_summary['QuarterStart_ts'] = pd.to_datetime(q_summary['QuarterStart'])
                        q_sorted = q_summary.sort_values(by='QuarterStart_ts')
                    elif 'Quarter' in q_summary.columns:
                        q_sorted = q_summary
                    else:
                        q_sorted = q_summary
                    for i in range(1, len(q_sorted)):
                        prev = q_sorted.iloc[i-1]
                        curr = q_sorted.iloc[i]
                        prev_label = prev.get('Quarter') or str(prev.get('QuarterStart'))
                        curr_label = curr.get('Quarter') or str(curr.get('QuarterStart'))
                        flux = curr.get('QoQ_Flux')
                        pct = curr.get('QoQ_Pct')
                        try:
                            flux_f = f"{float(flux):,.2f}"
                        except Exception:
                            flux_f = str(flux)
                        try:
                            pct_f = f"{float(pct):.1f}%"
                        except Exception:
                            pct_f = str(pct)
                        q_lines.append(f"{prev_label} -> {curr_label}: QoQ_Flux: {flux_f}, QoQ_Pct: {pct_f}")
                    if q_lines:
                        lines.append('\nExplicit adjacent-pair QoQ values:')
                        lines.append('\n'.join(q_lines))
                except Exception:
                    pass
        except Exception:
            pass

    memo_lines = []
    pair_lines = []
    quarter_memo_lines = []
    quarter_pair_lines = []
    dept_lines = []
    dept_pair_lines = []
    dept_memo_lines = []
    dept_memo_pair_lines = []
    class_lines = []
    class_pair_lines = []
    class_memo_lines = []
    class_memo_pair_lines = []
    if mode == 'mom' and months_sorted:
        # per-month top memos
        for period in months_sorted:
            group = month_groups[period]
            if memo_col and memo_col in group.columns:
                agg = group.groupby(memo_col)[value_col].sum().sort_values(ascending=False).head(10)
                if not agg.empty:
                    memo_lines.append(f"Top memos for {period.strftime('%Y-%m')}: ")
                    for memo_val, amt in agg.items():
                        memo_lines.append(f" - {memo_val}: {amt:.2f}")

            # per-month top departments
            if include_department and department_col in group.columns:
                try:
                    agg_dept = group.groupby(department_col)[value_col].sum().sort_values(ascending=False).head(10)
                    if not agg_dept.empty:
                        dept_lines.append(f"Top departments for {period.strftime('%Y-%m')}: ")
                        for dept_val, amt in agg_dept.items():
                            dept_lines.append(f" - {dept_val}: {amt:.2f}")
                except Exception:
                    pass

            # per-month top classes
            if include_class and class_col in group.columns:
                try:
                    agg_class = group.groupby(class_col)[value_col].sum().sort_values(ascending=False).head(10)
                    if not agg_class.empty:
                        class_lines.append(f"Top classes for {period.strftime('%Y-%m')}: ")
                        for class_val, amt in agg_class.items():
                            class_lines.append(f" - {class_val}: {amt:.2f}")
                except Exception:
                    pass

            # per-month top (Department, Memo) pairs
            if include_department and department_col in group.columns and memo_col and memo_col in group.columns:
                try:
                    agg_combo = group.groupby([department_col, memo_col])[value_col].sum().sort_values(ascending=False).head(10)
                    if not agg_combo.empty:
                        dept_memo_lines.append(f"Top (Department, Memo) for {period.strftime('%Y-%m')}: ")
                        for (dept_val, memo_val), amt in agg_combo.items():
                            dept_memo_lines.append(f" - ({dept_val}, {memo_val}): {amt:.2f}")
                except Exception:
                    pass

            # per-month top (Class, Memo) pairs
            if include_class and class_col in group.columns and memo_col and memo_col in group.columns:
                try:
                    agg_combo_c = group.groupby([class_col, memo_col])[value_col].sum().sort_values(ascending=False).head(10)
                    if not agg_combo_c.empty:
                        class_memo_lines.append(f"Top (Class, Memo) for {period.strftime('%Y-%m')}: ")
                        for (class_val, memo_val), amt in agg_combo_c.items():
                            class_memo_lines.append(f" - ({class_val}, {memo_val}): {amt:.2f}")
                except Exception:
                    pass

        # pairwise comparisons across all adjacent month pairs
        for i in range(1, len(months_sorted)):
            m_prev = months_sorted[i-1]
            m_curr = months_sorted[i]
            g_prev = month_groups[m_prev]
            g_curr = month_groups[m_curr]
            pair_lines.append(f"\nChanges {m_prev.strftime('%Y-%m')} -> {m_curr.strftime('%Y-%m')}: ")
            if memo_col and memo_col in g_prev.columns and memo_col in g_curr.columns:
                agg_prev = g_prev.groupby(memo_col)[value_col].sum()
                agg_curr = g_curr.groupby(memo_col)[value_col].sum()
                combined = pd.concat([agg_prev.rename('prev'), agg_curr.rename('curr')], axis=1).fillna(0)
                combined['delta'] = combined['curr'] - combined['prev']
                # safer percent: if prev is very small, mark pct as None/NaN to avoid huge misleading values
                with pd.option_context('mode.use_inf_as_na', True):
                    combined['pct'] = combined['delta'] / combined['prev'].replace({0: pd.NA})
                counts_prev = g_prev.groupby(memo_col)[value_col].count()
                counts_curr = g_curr.groupby(memo_col)[value_col].count()
                combined = combined.join(counts_prev.rename('count_prev'), how='left').join(counts_curr.rename('count_curr'), how='left').fillna(0)
                total_delta = combined['delta'].sum()
                if total_delta == 0:
                    combined['contrib_pct_of_flux'] = 0.0
                else:
                    combined['contrib_pct_of_flux'] = (combined['delta'] / total_delta) * 100
                top_pos = combined.sort_values(by='delta', ascending=False).head(5)
                top_neg = combined.sort_values(by='delta').head(5)
                def fmt_pct(v):
                    try:
                        if pd.isna(v):
                            return 'N/A'
                        # cap very large percent displays
                        if abs(v) > 1000:
                            return f">{1000:.0f}%"
                        return f"{v*100:.1f}%" if abs(v) >= 0.001 else f"{v*100:.2f}%"
                    except Exception:
                        return 'N/A'

                if not top_pos.empty:
                    pair_lines.append(' Top increases:')
                    for idx, row in top_pos.iterrows():
                        pct_display = fmt_pct(row.get('pct'))
                        pair_lines.append(f"  - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux', 0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f}) [count_prev={int(row.get('count_prev',0))} count_curr={int(row.get('count_curr',0))}] pct={pct_display}")
                if not top_neg.empty:
                    pair_lines.append(' Top decreases:')
                    for idx, row in top_neg.iterrows():
                        pct_display = fmt_pct(row.get('pct'))
                        pair_lines.append(f"  - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux', 0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f}) [count_prev={int(row.get('count_prev',0))} count_curr={int(row.get('count_curr',0))}] pct={pct_display}")
                new_memos = set(agg_curr.index) - set(agg_prev.index)
                gone_memos = set(agg_prev.index) - set(agg_curr.index)
                if new_memos:
                    pair_lines.append(f"  New memos in {m_curr.strftime('%Y-%m')}: {', '.join(list(new_memos)[:5])}")
                if gone_memos:
                    pair_lines.append(f"  Disappeared memos since {m_prev.strftime('%Y-%m')}: {', '.join(list(gone_memos)[:5])}")
                examples = []
                top_memos_examples = list(top_pos.index[:3]) + list(top_neg.index[:3])
                seen = set()
                for memo_val in top_memos_examples:
                    if memo_val in seen:
                        continue
                    seen.add(memo_val)
                    ex_prev = g_prev[g_prev[memo_col] == memo_val][[memo_col, value_col]].head(3)
                    ex_curr = g_curr[g_curr[memo_col] == memo_val][[memo_col, value_col]].head(3)
                    if not ex_prev.empty:
                        examples.append(f" Examples for '{memo_val}' in {m_prev.strftime('%Y-%m')}: ")
                        for _, r in ex_prev.iterrows():
                            examples.append(f"  - {r.get(memo_col)} : {r.get(value_col):.2f}")
                    if not ex_curr.empty:
                        examples.append(f" Examples for '{memo_val}' in {m_curr.strftime('%Y-%m')}: ")
                        for _, r in ex_curr.iterrows():
                            examples.append(f"  - {r.get(memo_col)} : {r.get(value_col):.2f}")
                if examples:
                    pair_lines.extend(examples)
            else:
                pair_lines.append('  Memo/description column not found; cannot compute memo-level diffs for this pair.')

            # Department-level pairwise diffs
            if include_department and department_col in g_prev.columns and department_col in g_curr.columns:
                try:
                    agg_prev_d = g_prev.groupby(department_col)[value_col].sum()
                    agg_curr_d = g_curr.groupby(department_col)[value_col].sum()
                    combined_d = pd.concat([agg_prev_d.rename('prev'), agg_curr_d.rename('curr')], axis=1).fillna(0)
                    combined_d['delta'] = combined_d['curr'] - combined_d['prev']
                    total_delta_d = combined_d['delta'].sum()
                    if total_delta_d == 0:
                        combined_d['contrib_pct_of_flux'] = 0.0
                    else:
                        combined_d['contrib_pct_of_flux'] = (combined_d['delta'] / total_delta_d) * 100
                    top_pos_d = combined_d.sort_values(by='delta', ascending=False).head(5)
                    top_neg_d = combined_d.sort_values(by='delta').head(5)
                    dept_pair_lines.append(' Department-level changes:')
                    if not top_pos_d.empty:
                        dept_pair_lines.append('  Top increases by department:')
                        for idx, row in top_pos_d.iterrows():
                            dept_pair_lines.append(f"   - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                    if not top_neg_d.empty:
                        dept_pair_lines.append('  Top decreases by department:')
                        for idx, row in top_neg_d.iterrows():
                            dept_pair_lines.append(f"   - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                except Exception:
                    pass

            # Class-level pairwise diffs
            if include_class and class_col in g_prev.columns and class_col in g_curr.columns:
                try:
                    agg_prev_cls = g_prev.groupby(class_col)[value_col].sum()
                    agg_curr_cls = g_curr.groupby(class_col)[value_col].sum()
                    combined_cls = pd.concat([agg_prev_cls.rename('prev'), agg_curr_cls.rename('curr')], axis=1).fillna(0)
                    combined_cls['delta'] = combined_cls['curr'] - combined_cls['prev']
                    total_delta_cls = combined_cls['delta'].sum()
                    if total_delta_cls == 0:
                        combined_cls['contrib_pct_of_flux'] = 0.0
                    else:
                        combined_cls['contrib_pct_of_flux'] = (combined_cls['delta'] / total_delta_cls) * 100
                    top_pos_cls = combined_cls.sort_values(by='delta', ascending=False).head(5)
                    top_neg_cls = combined_cls.sort_values(by='delta').head(5)
                    class_pair_lines.append(' Class-level changes:')
                    if not top_pos_cls.empty:
                        class_pair_lines.append('  Top increases by class:')
                        for idx, row in top_pos_cls.iterrows():
                            class_pair_lines.append(f"   - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                    if not top_neg_cls.empty:
                        class_pair_lines.append('  Top decreases by class:')
                        for idx, row in top_neg_cls.iterrows():
                            class_pair_lines.append(f"   - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                except Exception:
                    pass

            # (Department, Memo) pairwise diffs
            if include_department and department_col in g_prev.columns and department_col in g_curr.columns and memo_col and memo_col in g_prev.columns and memo_col in g_curr.columns:
                try:
                    agg_prev_c = g_prev.groupby([department_col, memo_col])[value_col].sum()
                    agg_curr_c = g_curr.groupby([department_col, memo_col])[value_col].sum()
                    combined_c = pd.concat([agg_prev_c.rename('prev'), agg_curr_c.rename('curr')], axis=1).fillna(0)
                    combined_c['delta'] = combined_c['curr'] - combined_c['prev']
                    total_delta_c = combined_c['delta'].sum()
                    if total_delta_c == 0:
                        combined_c['contrib_pct_of_flux'] = 0.0
                    else:
                        combined_c['contrib_pct_of_flux'] = (combined_c['delta'] / total_delta_c) * 100
                    top_pos_c = combined_c.sort_values(by='delta', ascending=False).head(5)
                    top_neg_c = combined_c.sort_values(by='delta').head(5)
                    dept_memo_pair_lines.append(' (Department, Memo) changes:')
                    if not top_pos_c.empty:
                        dept_memo_pair_lines.append('  Top increases by (Department, Memo):')
                        for idx, row in top_pos_c.iterrows():
                            dept_memo_pair_lines.append(f"   - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                    if not top_neg_c.empty:
                        dept_memo_pair_lines.append('  Top decreases by (Department, Memo):')
                        for idx, row in top_neg_c.iterrows():
                            dept_memo_pair_lines.append(f"   - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                except Exception:
                    pass

            # (Class, Memo) pairwise diffs
            if include_class and class_col in g_prev.columns and class_col in g_curr.columns and memo_col and memo_col in g_prev.columns and memo_col in g_curr.columns:
                try:
                    agg_prev_cm = g_prev.groupby([class_col, memo_col])[value_col].sum()
                    agg_curr_cm = g_curr.groupby([class_col, memo_col])[value_col].sum()
                    combined_cm = pd.concat([agg_prev_cm.rename('prev'), agg_curr_cm.rename('curr')], axis=1).fillna(0)
                    combined_cm['delta'] = combined_cm['curr'] - combined_cm['prev']
                    total_delta_cm = combined_cm['delta'].sum()
                    if total_delta_cm == 0:
                        combined_cm['contrib_pct_of_flux'] = 0.0
                    else:
                        combined_cm['contrib_pct_of_flux'] = (combined_cm['delta'] / total_delta_cm) * 100
                    top_pos_cm = combined_cm.sort_values(by='delta', ascending=False).head(5)
                    top_neg_cm = combined_cm.sort_values(by='delta').head(5)
                    class_memo_pair_lines.append(' (Class, Memo) changes:')
                    if not top_pos_cm.empty:
                        class_memo_pair_lines.append('  Top increases by (Class, Memo):')
                        for idx, row in top_pos_cm.iterrows():
                            class_memo_pair_lines.append(f"   - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                    if not top_neg_cm.empty:
                        class_memo_pair_lines.append('  Top decreases by (Class, Memo):')
                        for idx, row in top_neg_cm.iterrows():
                            class_memo_pair_lines.append(f"   - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                except Exception:
                    pass

    # Per-quarter memo analysis: if quarter_groups detected, perform analogous QoQ memo diffs
    if mode == 'qoq':
        try:
            if quarter_groups:
                # determine sorted quarter keys: if keys are timestamps use them, else use sheet order
                q_keys = list(quarter_groups.keys())
                # if keys are timestamps, sort; otherwise keep sheet order
                try:
                    q_sorted_keys = sorted([k for k in q_keys if isinstance(k, (pd.Timestamp, pd.DatetimeIndex))])
                except Exception:
                    q_sorted_keys = q_keys
                # if keys are strings (sheet names like FY2025-Q1), keep that order as found
                for i in range(1, len(q_keys)):
                    k_prev = q_keys[i-1]
                    k_curr = q_keys[i]
                    g_prev = quarter_groups.get(k_prev)
                    g_curr = quarter_groups.get(k_curr)
                    label_prev = str(k_prev)
                    label_curr = str(k_curr)
                    quarter_pair_lines.append(f"\nChanges {label_prev} -> {label_curr}: ")
                    if g_prev is None or g_curr is None:
                        quarter_pair_lines.append('  Quarter detail not available; skipped.')
                        continue
                    # detect memo column in quarter groups
                    q_memo_col = None
                    lowered = {c.lower(): c for c in g_prev.columns}
                    for name in possible_memo_names:
                        if name in g_prev.columns:
                            q_memo_col = name
                            break
                        if name.lower() in lowered:
                            q_memo_col = lowered[name.lower()]
                            break
                    if q_memo_col and q_memo_col in g_prev.columns and q_memo_col in g_curr.columns:
                        agg_prev = g_prev.groupby(q_memo_col)[value_col].sum()
                        agg_curr = g_curr.groupby(q_memo_col)[value_col].sum()
                        combined = pd.concat([agg_prev.rename('prev'), agg_curr.rename('curr')], axis=1).fillna(0)
                        combined['delta'] = combined['curr'] - combined['prev']
                        total_delta = combined['delta'].sum()
                        if total_delta == 0:
                            combined['contrib_pct_of_flux'] = 0.0
                        else:
                            combined['contrib_pct_of_flux'] = (combined['delta'] / total_delta) * 100
                        top_pos = combined.sort_values(by='delta', ascending=False).head(5)
                        top_neg = combined.sort_values(by='delta').head(5)
                        if not top_pos.empty:
                            quarter_pair_lines.append(' Top increases:')
                            for idx, row in top_pos.iterrows():
                                quarter_pair_lines.append(f"  - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                        if not top_neg.empty:
                            quarter_pair_lines.append(' Top decreases:')
                            for idx, row in top_neg.iterrows():
                                quarter_pair_lines.append(f"  - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux',0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f})")
                    else:
                        quarter_pair_lines.append('  Memo/description column not found; cannot compute QoQ memo-level diffs for this pair.')
        except Exception:
            pass

    if mode == 'mom':
        if pair_lines:
            lines.append('\n'.join(pair_lines))
        if memo_lines:
            lines.append('\n'.join(memo_lines))
        if include_department and dept_lines:
            lines.append('\n'.join(dept_lines))
        if include_department and dept_pair_lines:
            lines.append('\n'.join(dept_pair_lines))
        if include_department and dept_memo_lines:
            lines.append('\n'.join(dept_memo_lines))
        if include_department and dept_memo_pair_lines:
            lines.append('\n'.join(dept_memo_pair_lines))
        if include_class and class_lines:
            lines.append('\n'.join(class_lines))
        if include_class and class_pair_lines:
            lines.append('\n'.join(class_pair_lines))
        if include_class and class_memo_lines:
            lines.append('\n'.join(class_memo_lines))
        if include_class and class_memo_pair_lines:
            lines.append('\n'.join(class_memo_pair_lines))
    if mode == 'qoq':
        if quarter_pair_lines:
            lines.append('\n'.join(quarter_pair_lines))
        if quarter_memo_lines:
            lines.append('\n'.join(quarter_memo_lines))

    prompt = '\n'.join(lines)
    # Avoid truncating the prompt aggressively in code; rely on model max_tokens instead.
    # However, if the prompt is extremely long, keep a configurable safety cap.
    try:
        cap = int(prompt_cap) if prompt_cap is not None else 200000
    except Exception:
        cap = 200000
    if cap > 0 and len(prompt) > cap:
        prompt = prompt[:cap] + '\n...[truncated by tool]'

    default_model = 'gpt-5-2025-08-07'
    model = model or os.environ.get('OPENAI_MODEL') or default_model
    mode_label = 'month-over-month' if mode == 'mom' else 'quarter-over-quarter'
    print(f'Calling OpenAI model {model} to analyze {mode_label} fluctuations (this may incur usage). max_tokens={max_tokens or "unlimited"}')
    if dry_run:
        # return the prepared prompt for inspection without calling OpenAI
        return prompt

    try:
        base_url = os.environ.get('OPENAI_BASE_URL') or os.environ.get('SMG_BASE_URL')
        try:
            from openai import OpenAI as OpenAIClient
            try:
                if base_url:
                    client = OpenAIClient(api_key=key, base_url=base_url)
                else:
                    client = OpenAIClient(api_key=key)
            except TypeError:
                client = OpenAIClient(api_key=key)
                if base_url:
                    try:
                        client.base_url = base_url
                    except Exception:
                        pass
            create_kwargs = dict(model=model, messages=[{'role': 'system', 'content': 'You are a helpful financial analyst.'}, {'role': 'user', 'content': prompt}])
            if max_tokens is not None:
                create_kwargs['max_tokens'] = int(max_tokens)
            resp = client.chat.completions.create(**create_kwargs)
            # Extract content and detect truncation via finish_reason if available
            try:
                analysis = resp.choices[0].message.content.strip()
            except Exception:
                analysis = (getattr(resp.choices[0].message, 'content', '') or '')
            finish_reason = getattr(resp.choices[0], 'finish_reason', None)
            truncated_by_openai = (finish_reason == 'length')
        except Exception:
            import openai as legacy_openai
            if base_url:
                try:
                    legacy_openai.base_url = base_url
                except Exception:
                    pass
                try:
                    legacy_openai.api_base = base_url
                except Exception:
                    pass
            legacy_openai.api_key = key
            create_kwargs = dict(model=model, temperature=0.0, messages=[{'role': 'system', 'content': 'You are a helpful financial analyst.'}, {'role': 'user', 'content': prompt}])
            if max_tokens is not None:
                create_kwargs['max_tokens'] = int(max_tokens)
            resp = legacy_openai.chat.completions.create(**create_kwargs)
            try:
                analysis = resp['choices'][0]['message']['content'].strip()
            except Exception:
                analysis = getattr(resp.choices[0].message, 'content', '') or str(resp)
            # legacy response finish reason may live in resp['choices'][0].get('finish_reason')
            try:
                finish_reason = resp['choices'][0].get('finish_reason') if isinstance(resp, dict) else getattr(resp.choices[0], 'finish_reason', None)
            except Exception:
                finish_reason = None
            truncated_by_openai = (finish_reason == 'length')
    except Exception as e:
        print('OpenAI request failed:', e)
        return None

    # If OpenAI truncated the response due to token limits, append an explicit notice so it's obvious.
    if 'truncated_by_openai' in locals() and truncated_by_openai:
        analysis = (analysis or '') + '\n\n...[truncated by OpenAI due to max_tokens; consider increasing --openai-max-tokens or shortening prompt]'

    with open('openai_analysis.txt', 'w', encoding='utf-8') as f:
        f.write(analysis)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_excel)
        if 'AI_Analysis' in wb.sheetnames:
            ws = wb['AI_Analysis']
        else:
            ws = wb.create_sheet('AI_Analysis')
        ws.delete_cols(1, ws.max_column)
        for i, line in enumerate((analysis or '').splitlines(), start=1):
            ws.cell(row=i, column=1, value=line)
        wb.save(output_excel)
    except Exception:
        pass

    print('OpenAI analysis saved to openai_analysis.txt and appended to Excel (if possible).')
    return analysis


def _pause_before_exit():
    """Pause before exiting to allow user to see error messages in PyInstaller executable."""
    try:
        input("\nPress Enter to exit...")
    except (EOFError, KeyboardInterrupt):
        pass


def main():
    try:
        # Interactive prompts to collect inputs
        default_sheet = 0
        default_date_col = 'Period'
        default_value_col = 'Amount'
        output_excel = 'summary_flux.xlsx'

        print('Welcome to the Great P&L Flux Analyzer! This file is to analyze the flux of the P&L only, not the balance sheet or cash flow.')
        print('If you have issues or questions, please contact Ray Sang')
        
        try:
            input_path = input("Enter Excel/CSV filename to analyze: ").strip()
        except EOFError:
            input_path = ''
        if not input_path:
            print('No Excel filename provided. Aborting.')
            _pause_before_exit()
            return
        if not os.path.exists(input_path):
            print(f"Warning: '{input_path}' does not exist in {os.getcwd()}. We'll attempt to open it anyway.")

        # Ask whether to run AI analysis
        try:
            ai_choice = input('Run AI analysis? (y/N): ').strip().lower()
        except EOFError:
            ai_choice = ''
        should_analyze = ai_choice in ('y', 'yes', '1', 'true', 't')

        # Ask whether to refine by Department
        try:
            dept_choice = input("Analysis by Department? (y/N): ").strip().lower()
        except EOFError:
            dept_choice = ''
        include_department = dept_choice in ('y', 'yes', '1', 'true', 't')
        department_col = 'Department'

        # Ask whether to refine by Class
        try:
            class_choice = input("Analysis by Class? (y/N): ").strip().lower()
        except EOFError:
            class_choice = ''
        include_class = class_choice in ('y', 'yes', '1', 'true', 't')
        class_col = 'Class'

        # Ask for fiscal year end month (1-12). Example: 1 for January, 6 for June, etc.
        try:
            fy_end_input = input("Enter fiscal year end month (1-12). For Jan enter 1 [default 12]: ").strip()
        except EOFError:
            fy_end_input = ''
        try:
            fy_end_month = int(fy_end_input) if fy_end_input else 12
        except Exception:
            fy_end_month = 12
        if fy_end_month < 1 or fy_end_month > 12:
            print("Invalid fiscal year end month. Using 12 (December).")
            fy_end_month = 12

        analysis_mode = 'mom'
        if should_analyze:
            print("Choose analysis mode: [1] Month-over-Month (MoM), [2] Quarter-over-Quarter (QoQ)")
            try:
                mode_choice = input('Enter 1 for MoM or 2 for QoQ (default 1): ').strip()
            except EOFError:
                mode_choice = ''
            analysis_mode = 'qoq' if mode_choice == '2' else 'mom'

        try:
            summary = summarize(input_path, sheet=default_sheet, date_col=default_date_col, value_col=default_value_col, output_excel=output_excel, include_department=include_department, department_col=department_col, include_class=include_class, class_col=class_col, fy_end_month=fy_end_month)
        except Exception as e:
            print(f"Error: {e}")
            print('Aborting due to missing required columns. Please include Period, Amount, a Memo/Description column, and any selected optional columns (Department/Class).')
            _pause_before_exit()
            return
        print('Summary:')
        print(summary.to_string())
        print(f"Saved Excel to {output_excel}")

        if should_analyze:
            # Ensure OpenAI environment is configured before analysis
            ensure_openai_env()
            try:
                df_all = _read_table_file(input_path, sheet=default_sheet)
                # Normalize columns for AI path as well
                lowered_all = {str(c).strip().lower(): c for c in df_all.columns}
                if default_date_col.lower() in lowered_all and lowered_all[default_date_col.lower()] != default_date_col:
                    default_date_col = lowered_all[default_date_col.lower()]
                if default_value_col.lower() in lowered_all and lowered_all[default_value_col.lower()] != default_value_col:
                    default_value_col = lowered_all[default_value_col.lower()]
                _coerce_amount_inplace(df_all, default_value_col)
            except Exception as e:
                print('Failed to read input file for OpenAI analysis:', e)
                df_all = None
            if df_all is not None:
                perform_openai_analysis(output_excel, df_all, summary, model=None, max_tokens=None, prompt_cap=0, analysis_mode=analysis_mode, include_department=include_department, department_col=department_col, include_class=include_class, class_col=class_col)
        
        _pause_before_exit()
        
    except Exception as e:
        print(f"\nUnexpected error occurred: {e}")
        import traceback
        print("\nFull error details:")
        traceback.print_exc()
        _pause_before_exit()


if __name__ == '__main__':
    main()
