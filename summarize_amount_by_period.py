#!/usr/bin/env python3
"""
summarize_amount_by_period.py

Read an Excel file, summarize the `Amount` column by monthly `Period`,
compute month-over-month flux and percent change, save a summary Excel,
and create a month-over-month chart (Amount and MoM flux).
"""
import argparse
import os
import pandas as pd
import textwrap

# (No charting libraries required) keep environment tidy
mpl_config_dir = os.path.join(os.getcwd(), ".mplconfig")
os.makedirs(mpl_config_dir, exist_ok=True)
os.environ["MPLCONFIGDIR"] = mpl_config_dir


def summarize(input_file, sheet=0, date_col="Period", value_col="Amount", output_excel="summary_flux.xlsx"):
    df = pd.read_excel(input_file, sheet_name=sheet)
    if date_col not in df.columns:
        raise ValueError(f"Date/Period column '{date_col}' not found in input. Available: {list(df.columns)}")

    # Parse dates (coerce invalids)
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    if df[date_col].isna().all():
        # Try parsing as Period (like '2025-01' or '202501')
        try:
            df[date_col] = pd.PeriodIndex(df[date_col].astype(str), freq='M').to_timestamp()
        except Exception:
            raise ValueError(f"Could not parse '{date_col}' as dates or periods.")

    # Create monthly period (timestamp at month start)
    df['PeriodMonth'] = df[date_col].dt.to_period('M').dt.to_timestamp()

    if value_col not in df.columns:
        raise ValueError(f"Value column '{value_col}' not found in input. Available: {list(df.columns)}")

    # Aggregate
    monthly = df.groupby('PeriodMonth', sort=True)[value_col].sum().rename('Amount')
    flux = monthly.diff().fillna(0).rename('MoM_Flux')
    pct = monthly.pct_change().fillna(0).rename('MoM_Pct') * 100

    summary = pd.concat([monthly, flux, pct], axis=1)

    # Save to Excel: include a Summary sheet and one sheet per month with transactions
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # Summary sheet (trend) - write Period as YYYY-MM strings so Excel x-axis shows months
        summary_df = summary.reset_index()
        # Force Period column to YYYY-MM strings so Excel will use those as categorical x-axis labels
        try:
            summary_df.iloc[:, 0] = pd.to_datetime(summary_df.iloc[:, 0]).dt.strftime('%Y-%m')
        except Exception:
            # fallback: cast to string
            summary_df.iloc[:, 0] = summary_df.iloc[:, 0].astype(str)
        summary_df.rename(columns={summary_df.columns[0]: 'Period'}, inplace=True)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Per-month transaction sheets: filter original df for each month
        for period_ts, group in df.groupby('PeriodMonth'):
            # Sheet name like 'YYYY-MM'
            sheet_name = period_ts.strftime('%Y-%m')
            # write group's original columns (reset index)
            if 'PeriodMonth' in group.columns:
                group_sorted = group.sort_values(by=['PeriodMonth'], ignore_index=True)
            else:
                group_sorted = group.reset_index(drop=True)
            # Drop internal PeriodMonth column to keep original appearance
            group_to_write = group_sorted.drop(columns=['PeriodMonth'], errors='ignore')
            # Excel has a 31-char sheet name limit; ensure name fits
            sheet_name = sheet_name[:31]
            try:
                group_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception:
                # fallback: use a safe sheet name enumerated
                safe_name = sheet_name[:25] + '_' + str(abs(hash(sheet_name)) % 1000)
                group_to_write.to_excel(writer, sheet_name=safe_name, index=False)

    # Chart generation and embedding removed by user request.

    return summary


def perform_openai_analysis(output_excel, df_all, summary_df, api_env_var='OPENAI_API_KEY', date_col='Period', value_col='Amount', model=None, dry_run=False, max_tokens=None, prompt_cap=0):
    """Call OpenAI (if API key present in env) to analyze MoM fluctuations.

    This function prefers per-month sheets in `output_excel` (sheets named `YYYY-MM`) and
    analyzes all adjacent month pairs found there. If those sheets are missing it falls
    back to grouping `df_all` by `PeriodMonth`.
    """
    # If this is a dry run, skip loading .env and OpenAI imports/keys
    if not dry_run:
        # load .env if present
        env_path = os.path.join(os.getcwd(), '.env')
        if os.path.exists(env_path):
            try:
                with open(env_path, 'r') as f:
                    for ln in f:
                        ln = ln.strip()
                        if not ln or ln.startswith('#'):
                            continue
                        if '=' not in ln:
                            continue
                        k, v = ln.split('=', 1)
                        k = k.strip()
                        v = v.strip().strip('"').strip("'")
                        if k not in os.environ:
                            os.environ[k] = v
            except Exception:
                pass

        key = os.environ.get(api_env_var)
        if not key:
            print(f"OPENAI API key not found in environment variable '{api_env_var}'. Skipping OpenAI analysis.")
            return None

        try:
            import openai
        except Exception:
            print("openai package not installed. Install with: pip install openai")
            return None

        openai.api_key = key

    # Build initial prompt pieces
    lines = [
        "You are given monthly totals and transaction memo breakdowns. Explain the main month-over-month fluctuations and likely causes.",
        "Summary table (Period, Amount, MoM_Flux, MoM_Pct):",
        summary_df.to_string(index=False)
    ]

    # detect memo column
    memo_col = None
    possible_memo_names = ['memo', 'Memo', 'MEMO', 'description', 'Description', 'DESC', 'Details', 'notes', 'Notes']
    for c in df_all.columns:
        if c in possible_memo_names:
            memo_col = c
            break
    if memo_col is None:
        lowered = {c.lower(): c for c in df_all.columns}
        for name in possible_memo_names:
            if name.lower() in lowered:
                memo_col = lowered[name.lower()]
                break

    # read per-month sheets from output_excel if available
    month_groups = {}
    months_sorted = []
    try:
        xl = pd.ExcelFile(output_excel)
        import re
        month_sheet_names = [s for s in xl.sheet_names if re.match(r"^\d{4}-\d{2}$", s)]
        for s in month_sheet_names:
            try:
                df_sheet = pd.read_excel(output_excel, sheet_name=s)
                period_ts = pd.to_datetime(s + '-01')
                df_sheet['PeriodMonth'] = period_ts
                month_groups[period_ts] = df_sheet
            except Exception:
                continue
        months_sorted = sorted(month_groups.keys())
    except Exception:
        months_sorted = []

    # fallback to grouping raw dataframe
    if not months_sorted:
        if date_col not in df_all.columns:
            possible = [c for c in df_all.columns if 'period' in c.lower() or 'date' in c.lower()]
            if possible:
                date_col = possible[0]
        if date_col in df_all.columns:
            df_all[date_col] = pd.to_datetime(df_all[date_col], errors='coerce')
            df_all['PeriodMonth'] = df_all[date_col].dt.to_period('M').dt.to_timestamp()
        if 'PeriodMonth' in df_all.columns:
            month_groups = {period: grp for period, grp in df_all.groupby('PeriodMonth')}
            months_sorted = sorted(month_groups.keys())

    # Add an explicit, unambiguous mapping of adjacent month pairs to their MoM values
    # This prevents downstream consumers (or models) from mis-associating the MoM value
    # with the wrong month-pair when reading the summary table.
    try:
        pair_summary_lines = []
        # helper to extract MoM values from the passed summary_df for a given period (timestamp)
        def _get_mom_values_for(period_ts):
            mom_flux = None
            mom_pct = None
            try:
                # If summary_df has the period as its index
                if hasattr(summary_df, 'index') and period_ts in summary_df.index:
                    row = summary_df.loc[period_ts]
                    mom_flux = row.get('MoM_Flux') if 'MoM_Flux' in summary_df.columns else None
                    mom_pct = row.get('MoM_Pct') if 'MoM_Pct' in summary_df.columns else None
                else:
                    # try matching by YYYY-MM string in a 'Period' column or index string
                    key = period_ts.strftime('%Y-%m')
                    if 'Period' in summary_df.columns:
                        match = summary_df[summary_df['Period'].astype(str).str.startswith(key)]
                        if not match.empty:
                            mom_flux = match.iloc[0].get('MoM_Flux', None)
                            mom_pct = match.iloc[0].get('MoM_Pct', None)
                    else:
                        # attempt to find an index that starts with the key
                        for idx in summary_df.index:
                            if str(idx).startswith(key):
                                row = summary_df.loc[idx]
                                mom_flux = row.get('MoM_Flux') if 'MoM_Flux' in summary_df.columns else None
                                mom_pct = row.get('MoM_Pct') if 'MoM_Pct' in summary_df.columns else None
                                break
            except Exception:
                pass
            return mom_flux, mom_pct

        if months_sorted and len(months_sorted) > 1:
            for i in range(1, len(months_sorted)):
                m_prev = months_sorted[i-1]
                m_curr = months_sorted[i]
                mom_flux, mom_pct = _get_mom_values_for(m_curr)
                if mom_flux is None or (isinstance(mom_flux, float) and pd.isna(mom_flux)):
                    pair_summary_lines.append(f"{m_prev.strftime('%Y-%m')} -> {m_curr.strftime('%Y-%m')}: MoM_Flux: N/A, MoM_Pct: N/A")
                else:
                    # format numbers nicely
                    try:
                        mom_flux_f = f"{mom_flux:,.2f}"
                    except Exception:
                        mom_flux_f = str(mom_flux)
                    try:
                        mom_pct_f = f"{mom_pct:.1f}%" if mom_pct is not None and not pd.isna(mom_pct) else 'N/A'
                    except Exception:
                        mom_pct_f = str(mom_pct)
                    pair_summary_lines.append(f"{m_prev.strftime('%Y-%m')} -> {m_curr.strftime('%Y-%m')}: MoM_Flux: {mom_flux_f}, MoM_Pct: {mom_pct_f}")

        if pair_summary_lines:
            lines.append('\nExplicit adjacent-pair MoM values:')
            lines.append('\n'.join(pair_summary_lines))
    except Exception:
        # if anything goes wrong building the explicit mapping, continue without it
        pass

    memo_lines = []
    pair_lines = []
    if months_sorted:
        # per-month top memos
        for period in months_sorted:
            group = month_groups[period]
            if memo_col and memo_col in group.columns:
                agg = group.groupby(memo_col)[value_col].sum().sort_values(ascending=False).head(10)
                if not agg.empty:
                    memo_lines.append(f"Top memos for {period.strftime('%Y-%m')}: ")
                    for memo_val, amt in agg.items():
                        memo_lines.append(f" - {memo_val}: {amt:.2f}")

        # pairwise comparisons across all adjacent month pairs
        for i in range(1, len(months_sorted)):
            m_prev = months_sorted[i-1]
            m_curr = months_sorted[i]
            g_prev = month_groups[m_prev]
            g_curr = month_groups[m_curr]
            pair_lines.append(f"\nChanges {m_prev.strftime('%Y-%m')} -> {m_curr.strftime('%Y-%m')}: ")
            if memo_col and memo_col in g_prev.columns and memo_col in g_curr.columns:
                agg_prev = g_prev.groupby(memo_col)[value_col].sum()
                agg_curr = g_curr.groupby(memo_col)[value_col].sum()
                combined = pd.concat([agg_prev.rename('prev'), agg_curr.rename('curr')], axis=1).fillna(0)
                combined['delta'] = combined['curr'] - combined['prev']
                # safer percent: if prev is very small, mark pct as None/NaN to avoid huge misleading values
                with pd.option_context('mode.use_inf_as_na', True):
                    combined['pct'] = combined['delta'] / combined['prev'].replace({0: pd.NA})
                counts_prev = g_prev.groupby(memo_col)[value_col].count()
                counts_curr = g_curr.groupby(memo_col)[value_col].count()
                combined = combined.join(counts_prev.rename('count_prev'), how='left').join(counts_curr.rename('count_curr'), how='left').fillna(0)
                total_delta = combined['delta'].sum()
                if total_delta == 0:
                    combined['contrib_pct_of_flux'] = 0.0
                else:
                    combined['contrib_pct_of_flux'] = (combined['delta'] / total_delta) * 100
                top_pos = combined.sort_values(by='delta', ascending=False).head(5)
                top_neg = combined.sort_values(by='delta').head(5)
                def fmt_pct(v):
                    try:
                        if pd.isna(v):
                            return 'N/A'
                        # cap very large percent displays
                        if abs(v) > 1000:
                            return f">{1000:.0f}%"
                        return f"{v*100:.1f}%" if abs(v) >= 0.001 else f"{v*100:.2f}%"
                    except Exception:
                        return 'N/A'

                if not top_pos.empty:
                    pair_lines.append(' Top increases:')
                    for idx, row in top_pos.iterrows():
                        pct_display = fmt_pct(row.get('pct'))
                        pair_lines.append(f"  - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux', 0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f}) [count_prev={int(row.get('count_prev',0))} count_curr={int(row.get('count_curr',0))}] pct={pct_display}")
                if not top_neg.empty:
                    pair_lines.append(' Top decreases:')
                    for idx, row in top_neg.iterrows():
                        pct_display = fmt_pct(row.get('pct'))
                        pair_lines.append(f"  - {idx}: delta {row['delta']:.2f} ({row.get('contrib_pct_of_flux', 0.0):.1f}% of flux) (prev {row['prev']:.2f} -> curr {row['curr']:.2f}) [count_prev={int(row.get('count_prev',0))} count_curr={int(row.get('count_curr',0))}] pct={pct_display}")
                new_memos = set(agg_curr.index) - set(agg_prev.index)
                gone_memos = set(agg_prev.index) - set(agg_curr.index)
                if new_memos:
                    pair_lines.append(f"  New memos in {m_curr.strftime('%Y-%m')}: {', '.join(list(new_memos)[:5])}")
                if gone_memos:
                    pair_lines.append(f"  Disappeared memos since {m_prev.strftime('%Y-%m')}: {', '.join(list(gone_memos)[:5])}")
                examples = []
                top_memos_examples = list(top_pos.index[:3]) + list(top_neg.index[:3])
                seen = set()
                for memo_val in top_memos_examples:
                    if memo_val in seen:
                        continue
                    seen.add(memo_val)
                    ex_prev = g_prev[g_prev[memo_col] == memo_val][[memo_col, value_col]].head(3)
                    ex_curr = g_curr[g_curr[memo_col] == memo_val][[memo_col, value_col]].head(3)
                    if not ex_prev.empty:
                        examples.append(f" Examples for '{memo_val}' in {m_prev.strftime('%Y-%m')}: ")
                        for _, r in ex_prev.iterrows():
                            examples.append(f"  - {r.get(memo_col)} : {r.get(value_col):.2f}")
                    if not ex_curr.empty:
                        examples.append(f" Examples for '{memo_val}' in {m_curr.strftime('%Y-%m')}: ")
                        for _, r in ex_curr.iterrows():
                            examples.append(f"  - {r.get(memo_col)} : {r.get(value_col):.2f}")
                if examples:
                    pair_lines.extend(examples)
            else:
                pair_lines.append('  Memo/description column not found; cannot compute memo-level diffs for this pair.')

    else:
        memo_lines.append('No PeriodMonth could be computed from the input; skipping memo aggregates.')

    if pair_lines:
        lines.append('\n'.join(pair_lines))
    if memo_lines:
        lines.append('\n'.join(memo_lines))

    prompt = '\n'.join(lines)
    # Avoid truncating the prompt aggressively in code; rely on model max_tokens instead.
    # However, if the prompt is extremely long, keep a configurable safety cap.
    try:
        cap = int(prompt_cap) if prompt_cap is not None else 200000
    except Exception:
        cap = 200000
    if cap > 0 and len(prompt) > cap:
        prompt = prompt[:cap] + '\n...[truncated by tool]'

    default_model = 'gpt-4.1-2025-04-14'
    model = model or os.environ.get('OPENAI_MODEL') or default_model
    print(f'Calling OpenAI model {model} to analyze month-over-month fluctuations (this may incur usage). max_tokens={max_tokens or "unlimited"}')
    if dry_run:
        # return the prepared prompt for inspection without calling OpenAI
        return prompt

    try:
        base_url = os.environ.get('OPENAI_BASE_URL') or os.environ.get('SMG_BASE_URL')
        try:
            from openai import OpenAI as OpenAIClient
            try:
                if base_url:
                    client = OpenAIClient(api_key=key, base_url=base_url)
                else:
                    client = OpenAIClient(api_key=key)
            except TypeError:
                client = OpenAIClient(api_key=key)
                if base_url:
                    try:
                        client.base_url = base_url
                    except Exception:
                        pass
            create_kwargs = dict(model=model, messages=[{'role': 'system', 'content': 'You are a helpful financial analyst.'}, {'role': 'user', 'content': prompt}])
            if max_tokens is not None:
                create_kwargs['max_tokens'] = int(max_tokens)
            resp = client.chat.completions.create(**create_kwargs)
            # Extract content and detect truncation via finish_reason if available
            try:
                analysis = resp.choices[0].message.content.strip()
            except Exception:
                analysis = (getattr(resp.choices[0].message, 'content', '') or '')
            finish_reason = getattr(resp.choices[0], 'finish_reason', None)
            truncated_by_openai = (finish_reason == 'length')
        except Exception:
            import openai as legacy_openai
            if base_url:
                try:
                    legacy_openai.base_url = base_url
                except Exception:
                    pass
                try:
                    legacy_openai.api_base = base_url
                except Exception:
                    pass
            legacy_openai.api_key = key
            create_kwargs = dict(model=model, temperature=0.0, messages=[{'role': 'system', 'content': 'You are a helpful financial analyst.'}, {'role': 'user', 'content': prompt}])
            if max_tokens is not None:
                create_kwargs['max_tokens'] = int(max_tokens)
            resp = legacy_openai.chat.completions.create(**create_kwargs)
            try:
                analysis = resp['choices'][0]['message']['content'].strip()
            except Exception:
                analysis = getattr(resp.choices[0].message, 'content', '') or str(resp)
            # legacy response finish reason may live in resp['choices'][0].get('finish_reason')
            try:
                finish_reason = resp['choices'][0].get('finish_reason') if isinstance(resp, dict) else getattr(resp.choices[0], 'finish_reason', None)
            except Exception:
                finish_reason = None
            truncated_by_openai = (finish_reason == 'length')
    except Exception as e:
        print('OpenAI request failed:', e)
        return None

    # If OpenAI truncated the response due to token limits, append an explicit notice so it's obvious.
    if 'truncated_by_openai' in locals() and truncated_by_openai:
        analysis = (analysis or '') + '\n\n...[truncated by OpenAI due to max_tokens; consider increasing --openai-max-tokens or shortening prompt]'

    with open('openai_analysis.txt', 'w') as f:
        f.write(analysis)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_excel)
        if 'AI_Analysis' in wb.sheetnames:
            ws = wb['AI_Analysis']
        else:
            ws = wb.create_sheet('AI_Analysis')
        ws.delete_cols(1, ws.max_column)
        for i, line in enumerate((analysis or '').splitlines(), start=1):
            ws.cell(row=i, column=1, value=line)
        wb.save(output_excel)
    except Exception:
        pass

    print('OpenAI analysis saved to openai_analysis.txt and appended to Excel (if possible).')
    return analysis


def main():
    parser = argparse.ArgumentParser(description='Summarize Amount by Period')
    parser.add_argument('input', nargs='?', default='68100 details.xlsx', help='Input Excel file (default: 68100 details.xlsx)')
    parser.add_argument('--sheet', default=0, help='Sheet name or index (default: 0)')
    parser.add_argument('--date-col', default='Period', help="Date/period column name (default: 'Period')")
    parser.add_argument('--value-col', default='Amount', help="Value column to sum (default: 'Amount')")
    parser.add_argument('--output-excel', default='summary_flux.xlsx', help='Output Excel filename')
    parser.add_argument('--openai-analyze', action='store_true', help='If set, call OpenAI to analyze MoM fluctuations (requires OPENAI_API_KEY in env)')
    parser.add_argument('--openai-model', default=None, help='OpenAI model to use (default: gpt-4.1-2025-04-14)')
    parser.add_argument('--openai-max-tokens', default=None, help='Max tokens to request from OpenAI (default: unlimited). Provide an integer value, or leave empty for no limit.')
    parser.add_argument('--openai-prompt-cap', default=0, type=int, help='Maximum prompt length in characters before truncation (default: 0 meaning no client-side cap).')
    args = parser.parse_args()

    summary = summarize(args.input, sheet=args.sheet, date_col=args.date_col, value_col=args.value_col, output_excel=args.output_excel)
    print('Summary:')
    print(summary.to_string())
    print(f"Saved Excel to {args.output_excel}")

    # Decide whether to run OpenAI analysis: user flag OR env var presence
    should_analyze = getattr(args, 'openai_analyze', False) or (os.environ.get('OPENAI_API_KEY') is not None)
    if should_analyze:
        # read back the full data to pass memos and amounts
        try:
            df_all = pd.read_excel(args.input, sheet_name=args.sheet)
        except Exception as e:
            print('Failed to read input file for OpenAI analysis:', e)
            df_all = None
        if df_all is not None:
            # normalize max_tokens: if provided and numeric, use int, else None (no limit)
            mt = None
            if args.openai_max_tokens is not None and str(args.openai_max_tokens).strip() != '':
                try:
                    mt = int(args.openai_max_tokens)
                except Exception:
                    mt = None
            perform_openai_analysis(args.output_excel, df_all, summary, model=args.openai_model, max_tokens=mt, prompt_cap=args.openai_prompt_cap)


if __name__ == '__main__':
    main()
